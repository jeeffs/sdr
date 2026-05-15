import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import traceback

OUTPUT_PATH = r"C:\Users\Jeeff\OneDrive\Documentos\ObsidianAI\Claude\SDR\Relatorio_Analise_SDR.xlsx"

FILES = {
    "Juliano_Marco": r"C:\Users\Jeeff\OneDrive\Documentos\ObsidianAI\Claude\SDR\DADOS\planilhas\Juliano_Marco_2026.xlsx",
    "Prestadores":   r"C:\Users\Jeeff\OneDrive\Documentos\ObsidianAI\Claude\SDR\DADOS\planilhas\PRESTADORES.xlsx",
    "Jefferson":     r"C:\Users\Jeeff\OneDrive\Documentos\ObsidianAI\Claude\SDR\DADOS\planilhas\Planilha Simples - JEFFERSON RODRIGUES.xlsx",
    "Fabricacao":    r"C:\Users\Jeeff\OneDrive\Documentos\ObsidianAI\Claude\SDR\congresso-2026\Fabricacao_Camisetas_Congresso2026.xlsx",
    "Import_Lista":  r"C:\Users\Jeeff\OneDrive\Documentos\ObsidianAI\Claude\SDR\congresso-2026\ADV_Camisetas_Import_Lista.xlsx",
    "Estoque_Import":r"C:\Users\Jeeff\OneDrive\Documentos\ObsidianAI\Claude\SDR\congresso-2026\ADV_Estoque_Import.xlsx",
    "Entrega_Status":r"C:\Users\Jeeff\OneDrive\Documentos\ObsidianAI\Claude\SDR\congresso-2026\ADV_Entrega_Status.xlsx",
    "Faltantes":     r"C:\Users\Jeeff\OneDrive\Documentos\ObsidianAI\Claude\SDR\congresso-2026\Faltantes_Fabricacao_03042026.xlsx",
}

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
ALT_FILL     = PatternFill("solid", fgColor="DEEAF1")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT    = Font(name="Arial", size=10)
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SIZE_KEYWORDS = ["pp","p","m","g","gg","xg","xxg","2g","3g","infantil","baby","adulto","unico","tamanho","tam","size"]


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def style_body(cell, alt=False):
    cell.font = BODY_FONT
    cell.fill = ALT_FILL if alt else WHITE_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=False)
    cell.border = THIN_BORDER


def autofit(ws, min_w=15):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 3, 50))


def write_table(ws, headers, rows, start_row=1):
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        style_header(cell)
    for ri, row in enumerate(rows, start_row + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            style_body(cell, alt=(ri % 2 == 0))
    return start_row + len(rows) + 2


def detect_size_columns(df):
    size_cols = []
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if any(k == col_lower or col_lower.startswith(k) for k in ["pp","p","m","g","gg","xg","xxg","2g","3g"]):
            size_cols.append(col)
        elif any(k in col_lower for k in ["tamanho","tam ","size","infantil","baby"]):
            size_cols.append(col)
    return size_cols


def detect_monetary_columns(df):
    money_cols = []
    keywords = ["valor","total","preco","preco","custo","receita","pagamento","price","amount","r$","$"]
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if any(k in col_lower for k in keywords):
            if pd.api.types.is_numeric_dtype(df[col]):
                money_cols.append(col)
    return money_cols


def detect_status_columns(df):
    status_cols = []
    keywords = ["status","situacao","situação","estado","entrega","fase","etapa","tipo","situac"]
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if any(k in col_lower for k in keywords):
            if df[col].dtype == object or str(df[col].dtype) == "category":
                status_cols.append(col)
    return status_cols


def analyze_file(key, path):
    result = {"key": key, "path": path, "error": None, "sheets_data": {}, "summary": {}}
    try:
        sheets = pd.read_excel(path, sheet_name=None, dtype=str)
        result["summary"]["num_sheets"] = len(sheets)
        result["summary"]["sheet_names"] = list(sheets.keys())
        total_rows = 0
        for sname, df in sheets.items():
            # try to convert numeric columns
            for col in df.columns:
                try:
                    converted = pd.to_numeric(df[col], errors='coerce')
                    if converted.notna().sum() > df[col].notna().sum() * 0.3:
                        df[col] = converted
                except Exception:
                    pass
            total_rows += len(df)
            col_info = []
            for col in df.columns:
                non_null = df[col].notna().sum()
                pct = round(non_null / len(df) * 100, 1) if len(df) > 0 else 0
                col_info.append({"col": col, "dtype": str(df[col].dtype), "non_null": non_null, "total": len(df), "pct": pct})
            result["sheets_data"][sname] = {
                "df": df,
                "col_info": col_info,
                "size_cols": detect_size_columns(df),
                "money_cols": detect_monetary_columns(df),
                "status_cols": detect_status_columns(df),
            }
        result["summary"]["total_rows"] = total_rows
        all_cols = sum([len(sd["col_info"]) for sd in result["sheets_data"].values()])
        result["summary"]["total_cols"] = all_cols
        # overall fill rate
        all_non_null = sum([sum(c["non_null"] for c in sd["col_info"]) for sd in result["sheets_data"].values()])
        all_total = sum([sum(c["total"] for c in sd["col_info"]) for sd in result["sheets_data"].values()])
        result["summary"]["fill_pct"] = round(all_non_null / all_total * 100, 1) if all_total > 0 else 0
    except Exception as e:
        result["error"] = f"{type(e).__name__}: {e}\n{traceback.format_exc()}"
    return result


def section_title(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    cell.fill = PatternFill("solid", fgColor="E2EFDA")
    cell.alignment = Alignment(vertical="center")
    return row + 1


def build_detail_sheet(wb, sheet_name, analysis):
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"
    row = 1

    if analysis["error"]:
        ws.cell(row=row, column=1, value="ERRO AO LER ARQUIVO").font = Font(name="Arial", size=12, bold=True, color="FF0000")
        ws.cell(row=row+1, column=1, value=str(analysis["error"][:500]))
        autofit(ws)
        return

    for sname, sdata in analysis["sheets_data"].items():
        df = sdata["df"]

        row = section_title(ws, row, f"ABA: {sname}  |  {len(df)} linhas  x  {len(df.columns)} colunas")

        # Column info table
        headers = ["Coluna", "Tipo", "Preenchido", "Total Linhas", "% Preenchimento"]
        rows_data = [[c["col"], c["dtype"], c["non_null"], c["total"], f"{c['pct']}%"] for c in sdata["col_info"]]
        row = write_table(ws, headers, rows_data, start_row=row)

        # Monetary totals
        if sdata["money_cols"]:
            row = section_title(ws, row, "TOTAIS MONETARIOS")
            headers_m = ["Coluna", "Soma Total (R$)", "Media", "Minimo", "Maximo"]
            rows_m = []
            for col in sdata["money_cols"]:
                numeric = pd.to_numeric(df[col], errors="coerce")
                rows_m.append([col, round(numeric.sum(), 2), round(numeric.mean(), 2),
                                round(numeric.min(), 2), round(numeric.max(), 2)])
            row = write_table(ws, headers_m, rows_m, start_row=row)

        # Status counts
        for scol in sdata["status_cols"]:
            row = section_title(ws, row, f"CONTAGEM POR STATUS: {scol}")
            counts = df[scol].value_counts(dropna=False)
            headers_s = ["Valor", "Contagem", "% do Total"]
            rows_s = [[str(v), int(c), f"{round(c/len(df)*100,1)}%"] for v, c in counts.items()]
            row = write_table(ws, headers_s, rows_s, start_row=row)

        # Size totals
        if sdata["size_cols"]:
            row = section_title(ws, row, "TOTAIS POR TAMANHO")
            headers_sz = ["Coluna / Tamanho", "Soma", "Media", "Nao Nulo"]
            rows_sz = []
            for col in sdata["size_cols"]:
                numeric = pd.to_numeric(df[col], errors="coerce")
                if numeric.notna().sum() > 0:
                    rows_sz.append([col, int(numeric.sum()), round(numeric.mean(), 2), int(numeric.notna().sum())])
                else:
                    for v, c in df[col].value_counts(dropna=False).items():
                        rows_sz.append([f"{col} = {v}", c, "-", c])
            if rows_sz:
                row = write_table(ws, headers_sz, rows_sz, start_row=row)

        # Other numeric stats
        num_cols = df.select_dtypes(include="number").columns.tolist()
        extra_num = [c for c in num_cols if c not in sdata["money_cols"] and c not in sdata["size_cols"]]
        if extra_num:
            row = section_title(ws, row, "OUTRAS COLUNAS NUMERICAS")
            headers_n = ["Coluna", "Soma", "Media", "Min", "Max", "Nao Nulo"]
            rows_n = []
            for col in extra_num[:25]:
                rows_n.append([col, round(float(df[col].sum()), 2), round(float(df[col].mean()), 2),
                                round(float(df[col].min()), 2), round(float(df[col].max()), 2),
                                int(df[col].notna().sum())])
            row = write_table(ws, headers_n, rows_n, start_row=row)

        row += 1

    autofit(ws)


def build_resumo(wb, analyses):
    ws = wb.active
    ws.title = "Resumo_Executivo"
    ws.freeze_panes = "A2"

    headers = ["Chave", "Arquivo", "Status", "Abas", "Total Linhas", "Total Colunas", "% Preenchimento", "Observacoes"]
    rows_data = []
    for key, an in analyses.items():
        if an["error"]:
            rows_data.append([key, os.path.basename(an["path"]), "ERRO", "-", "-", "-", "-", an["error"][:200]])
        else:
            rows_data.append([
                key,
                os.path.basename(an["path"]),
                "OK",
                ", ".join(an["summary"].get("sheet_names", [])),
                an["summary"].get("total_rows", 0),
                an["summary"].get("total_cols", 0),
                f"{an['summary'].get('fill_pct', 0)}%",
                f"{an['summary'].get('num_sheets', 0)} aba(s)",
            ])

    write_table(ws, headers, rows_data, start_row=1)

    row = len(rows_data) + 4
    cell = ws.cell(row=row, column=1, value="METRICAS PRINCIPAIS POR ARQUIVO")
    cell.font = Font(name="Arial", size=12, bold=True, color="1F4E79")
    row += 2

    for key, an in analyses.items():
        if an["error"]:
            ws.cell(row=row, column=1, value=f"[ {key} ] - ERRO").font = Font(name="Arial", size=10, bold=True, color="FF0000")
            row += 1
            continue
        ws.cell(row=row, column=1, value=f"[ {key} ]").font = Font(name="Arial", size=10, bold=True, color="1F4E79")
        row += 1
        for sname, sdata in an["sheets_data"].items():
            df = sdata["df"]
            for col in sdata["money_cols"]:
                numeric = pd.to_numeric(df[col], errors="coerce")
                ws.cell(row=row, column=1, value=f"  {sname} > {col}: TOTAL = R$ {numeric.sum():,.2f}").font = BODY_FONT
                row += 1
            for col in sdata["size_cols"]:
                numeric = pd.to_numeric(df[col], errors="coerce")
                if numeric.notna().sum() > 0:
                    ws.cell(row=row, column=1, value=f"  {sname} > {col} [tamanho]: SOMA = {int(numeric.sum())}").font = BODY_FONT
                    row += 1
            for col in sdata["status_cols"]:
                counts = df[col].value_counts(dropna=False).head(6)
                summary = " | ".join([f"{v}: {c}" for v, c in counts.items()])
                ws.cell(row=row, column=1, value=f"  {sname} > [{col}]: {summary}").font = BODY_FONT
                row += 1
        row += 1

    autofit(ws)


def main():
    print("=" * 60)
    print("GERANDO RELATORIO SDR - ANALISE EXCEL")
    print("=" * 60)

    analyses = {}
    for key, path in FILES.items():
        print(f"\nLendo: {key}")
        print(f"  Arquivo: {path}")
        analyses[key] = analyze_file(key, path)
        if analyses[key]["error"]:
            print(f"  ERRO: {analyses[key]['error'][:120]}")
        else:
            s = analyses[key]["summary"]
            print(f"  OK: {s['num_sheets']} aba(s), {s['total_rows']} linhas, {s['total_cols']} colunas, {s['fill_pct']}% preenchido")

    print("\nCriando workbook Excel...")
    wb = openpyxl.Workbook()

    print("  Aba: Resumo_Executivo")
    build_resumo(wb, analyses)

    for key in FILES.keys():
        print(f"  Aba: {key}")
        build_detail_sheet(wb, key, analyses[key])

    print(f"\nSalvando em:\n  {OUTPUT_PATH}")
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print("\nArquivo salvo com sucesso!")

    print("\n" + "=" * 60)
    print("RESUMO FINAL")
    print("=" * 60)
    for key, an in analyses.items():
        if an["error"]:
            print(f"\n[ERRO] {key}: {an['error'][:150]}")
        else:
            s = an["summary"]
            print(f"\n[OK] {key} - {s['total_rows']} linhas | abas: {s['sheet_names']}")
            for sname, sdata in an["sheets_data"].items():
                df = sdata["df"]
                for col in sdata["money_cols"]:
                    numeric = pd.to_numeric(df[col], errors="coerce")
                    print(f"  {sname} > {col}: R$ {numeric.sum():,.2f}")
                for col in sdata["size_cols"]:
                    numeric = pd.to_numeric(df[col], errors="coerce")
                    if numeric.notna().sum() > 0:
                        print(f"  {sname} > {col} (tamanho): {int(numeric.sum())} unidades")
                for col in sdata["status_cols"]:
                    counts = df[col].value_counts(dropna=False).head(5)
                    print(f"  {sname} > status [{col}]: {dict(counts)}")


if __name__ == "__main__":
    main()
