"""
╔══════════════════════════════════════════════════════════════╗
║   VALIDADOR DE PLANILHAS — SISTEMA DE PRESTADORES           ║
║   4 Frentes × 3 Tabelas de Preço (V1/V2/V3)                ║
╚══════════════════════════════════════════════════════════════╝

Estrutura de pasta esperada:
  [Ano] > [Mes] >
      ├── PRESTADOR_FILIAL_MANUTENÇÃO.xlsm      (V1 — solto)
      ├── PRESTADOR_FILIAL_PROJETO.xlsm
      ├── PRESTADOR_MATRIZ_MANUTENÇÃO.xlsm
      ├── PRESTADOR_MATRIZ_PROJETO.xlsm
      ├── V2/
      │    ├── PRESTADOR_FILIAL_MANUTENÇÃO.xlsm  (V2 — supervisor)
      │    └── ...
      └── V3/
           ├── Projeto/
           │    ├── PRESTADOR_FILIAL_MANUTENÇÃO.xlsm  (V3 — admin)
           │    └── ...
           └── Geral/  <- IGNORADO
"""

import os
import re
import sys
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    messagebox.showerror("Dependência faltando",
        "Instale openpyxl:\n\npip install openpyxl")
    sys.exit(1)

# Módulo de referência (planilha IMPOSTO) — opcional
try:
    _this_dir = os.path.dirname(os.path.abspath(__file__))
    if _this_dir not in sys.path:
        sys.path.insert(0, _this_dir)
    from referencia import comparar_com_referencia as _comparar_ref
    _REF_DISPONIVEL = True
except Exception:
    _REF_DISPONIVEL = False
    def _comparar_ref(*a, **kw):
        return {}

# ══════════════════════════════════════════════════════════════
#  CONSTANTES
# ══════════════════════════════════════════════════════════════

FRENTES = [
    "FILIAL_MANUTENÇÃO",
    "FILIAL_PROJETO",
    "MATRIZ_MANUTENÇÃO",
    "MATRIZ_PROJETO",
]

FRENTE_LABELS = {
    "FILIAL_MANUTENÇÃO":  "🔧 Filial — Manutenção",
    "FILIAL_PROJETO":     "📐 Filial — Projeto",
    "MATRIZ_MANUTENÇÃO":  "🔧 Matriz — Manutenção",
    "MATRIZ_PROJETO":     "📐 Matriz — Projeto",
}

VERSION_LABELS = {
    "V1": "V1 · Prestador",
    "V2": "V2 · Supervisor",
    "V3": "V3 · Admin",
}

EXCEL_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}

# Cores
C_BG       = "#1E1E2E"
C_PANEL    = "#2A2A3E"
C_CARD     = "#313145"
C_ACCENT   = "#7C83FD"
C_GREEN    = "#50FA7B"
C_RED      = "#FF5555"
C_ORANGE   = "#FFB86C"
C_YELLOW   = "#F1FA8C"
C_TEXT     = "#F8F8F2"
C_MUTED    = "#9999BB"
C_BORDER   = "#44446A"

# ══════════════════════════════════════════════════════════════
#  MOTOR DE VALIDAÇÃO
# ══════════════════════════════════════════════════════════════

def normalize(text: str) -> str:
    """Remove espaços extras, normaliza para comparação."""
    if not text:
        return ""
    return " ".join(str(text).strip().split())


def load_depara(ws_depara) -> dict:
    """Carrega tabela de preços da aba de_para → {serviço: preço}."""
    prices = {}
    for row in ws_depara.iter_rows(min_row=1, values_only=True):
        srv, price = row[0], row[1]
        if srv and price is not None and isinstance(price, (int, float)):
            prices[normalize(srv)] = float(price)
    return prices


def get_columns(ws_base):
    """Detecta estrutura de colunas (com ou sem coluna CÓDIGO)."""
    header = [ws_base.cell(2, c).value for c in range(1, ws_base.max_column + 1)]
    has_codigo = "CÓDIGO" in header
    if has_codigo:
        return {
            "servicos": [(7,8,9),(10,11,12),(13,14,15),(16,17,18),(19,20,21)],
            "total": 22, "data_start": 3
        }
    else:
        return {
            "servicos": [(6,7,8),(9,10,11),(12,13,14),(15,16,17),(18,19,20)],
            "total": 21, "data_start": 3
        }


def validate_single_file(filepath: str) -> dict:
    """
    Valida um arquivo .xlsm.
    Retorna:
      {
        "ok": bool,
        "total_geral": float,
        "erros": [{"linha": int, "celula": str, "tipo": str,
                   "encontrado": str, "esperado": str, "como_corrigir": str}]
      }
    """
    erros = []
    total_geral = 0.0

    try:
        wb_f = openpyxl.load_workbook(filepath, keep_vba=True)
        wb_v = openpyxl.load_workbook(filepath, keep_vba=True, data_only=True)
    except Exception as e:
        return {"ok": False, "total_geral": 0.0, "erros": [{
            "linha": "-", "celula": "-",
            "tipo": "ERRO AO ABRIR ARQUIVO",
            "encontrado": str(e), "esperado": "Arquivo válido",
            "como_corrigir": "Verifique se o arquivo não está corrompido ou aberto em outro programa."
        }]}

    ws_f = wb_f["base"]
    ws_v = wb_v["base"]
    ws_dep = wb_v["de_para"]

    prices = load_depara(ws_dep)
    cols   = get_columns(ws_v)

    # Linha 1 = total geral
    tg_cell = ws_v.cell(1, ws_v.max_column)
    if isinstance(tg_cell.value, (int, float)):
        total_geral = tg_cell.value

    for row_idx in range(cols["data_start"], ws_v.max_row + 1):
        row_vals = [ws_v.cell(row_idx, c).value
                    for c in range(1, ws_v.max_column + 1)]
        if all(v is None for v in row_vals):
            continue

        soma = 0.0

        for s_col, q_col, v_col in cols["servicos"]:
            srv_v  = ws_v.cell(row_idx, s_col).value
            qtd_v  = ws_v.cell(row_idx, q_col).value
            val_v  = ws_v.cell(row_idx, v_col).value
            val_f  = ws_f.cell(row_idx, v_col).value
            cel    = f"{get_column_letter(v_col)}{row_idx}"

            # Erro Excel
            if isinstance(val_v, str) and val_v in EXCEL_ERRORS:
                erros.append({
                    "linha": row_idx, "celula": cel,
                    "tipo": "❌ Erro de fórmula Excel",
                    "encontrado": val_v, "esperado": "Valor numérico",
                    "como_corrigir":
                        f"Célula {cel} contém '{val_v}'. "
                        "Verifique se a referência da fórmula VLOOKUP está correta "
                        "e se a aba 'de_para' existe e tem os dados esperados."
                })
                continue

            val_num = val_v if isinstance(val_v, (int, float)) else 0.0

            if srv_v and normalize(str(srv_v)):
                srv_norm = normalize(str(srv_v))
                preco    = prices.get(srv_norm)

                # Serviço não encontrado
                if preco is None:
                    erros.append({
                        "linha": row_idx,
                        "celula": f"{get_column_letter(s_col)}{row_idx}",
                        "tipo": "⚠️ Serviço não encontrado na tabela",
                        "encontrado": srv_norm,
                        "esperado": "Serviço existente na aba de_para",
                        "como_corrigir":
                            f"O serviço '{srv_norm}' não existe na aba 'de_para'. "
                            "Adicione o serviço na tabela de preços ou corrija o nome "
                            "(verifique espaços e acentos)."
                    })
                    soma += val_num
                    continue

                qtd_num  = qtd_v if isinstance(qtd_v, (int, float)) else 0.0
                esperado = round(preco * qtd_num, 4)
                real     = round(val_num, 4)

                # Valor hardcoded
                is_hard = (isinstance(val_f, (int, float)) or
                           (isinstance(val_f, str) and not val_f.startswith("=")))
                if is_hard:
                    diff = round(abs(esperado - real), 2)
                    erros.append({
                        "linha": row_idx, "celula": cel,
                        "tipo": "🔒 Valor hardcoded (sem VLOOKUP)",
                        "encontrado": f"R${real} (fixo)",
                        "esperado": f"R${esperado} (via fórmula)",
                        "como_corrigir":
                            f"Célula {cel} tem valor R${real} digitado manualmente. "
                            f"Substitua por: =IF({get_column_letter(s_col)}{row_idx}<>\"\","
                            f"VLOOKUP({get_column_letter(s_col)}{row_idx},de_para!$A:$B,2,0)"
                            f"*{get_column_letter(q_col)}{row_idx},0)"
                            + (f"\n↳ Diferença de R${diff}" if diff > 0 else "")
                    })
                elif abs(esperado - real) > 0.01:
                    erros.append({
                        "linha": row_idx, "celula": cel,
                        "tipo": "🔢 Valor divergente (QTD × Preço)",
                        "encontrado": f"R${real}",
                        "esperado":   f"R${esperado} ({qtd_num} × R${preco})",
                        "como_corrigir":
                            f"O valor R${real} na célula {cel} não bate com "
                            f"{qtd_num} × R${preco} = R${esperado}. "
                            f"Corrija a quantidade em {get_column_letter(q_col)}{row_idx} "
                            f"ou o preço de '{srv_norm}' na aba de_para."
                    })

                soma += real
            else:
                if isinstance(val_v, (int, float)) and abs(val_v) > 0.01:
                    erros.append({
                        "linha": row_idx, "celula": cel,
                        "tipo": "🔢 Valor sem serviço associado",
                        "encontrado": f"R${val_v}",
                        "esperado": "R$0 (sem serviço)",
                        "como_corrigir":
                            f"A célula {cel} tem valor R${val_v} mas não há "
                            f"serviço na coluna {get_column_letter(s_col)}{row_idx}. "
                            "Limpe o valor ou preencha o serviço correspondente."
                    })

        # Total
        t_col  = cols["total"]
        tot_v  = ws_v.cell(row_idx, t_col).value
        if isinstance(tot_v, (int, float)):
            tot_real = round(tot_v, 4)
            soma_esp = round(soma, 4)
            if abs(tot_real - soma_esp) > 0.01:
                erros.append({
                    "linha": row_idx,
                    "celula": f"{get_column_letter(t_col)}{row_idx}",
                    "tipo": "➕ Total divergente",
                    "encontrado": f"R${tot_real}",
                    "esperado":   f"R${soma_esp}",
                    "como_corrigir":
                        f"O TOTAL da linha {row_idx} é R${tot_real} mas a soma dos "
                        f"valores é R${soma_esp}. "
                        f"Corrija a fórmula na coluna {get_column_letter(t_col)}: "
                        f"deve somar todas as colunas de VALOR."
                })

    wb_f.close()
    wb_v.close()
    return {"ok": len(erros) == 0, "total_geral": total_geral, "erros": erros}


def scan_folder(base_path: str) -> dict:
    """
    Escaneia a pasta do mês e retorna:
    {
      "providers": ["ORLANDO", "CARLOS", ...],
      "files": {
        "ORLANDO": {
          "V1": {"FILIAL_MANUTENÇÃO": "caminho.xlsm", ...},
          "V2": {...},
          "V3": {...}
        }
      },
      "warnings": [...]
    }
    """
    result   = {"providers": [], "files": {}, "warnings": []}
    frente_patterns = {
        "FILIAL_MANUTENÇÃO":  re.compile(r"FILIAL[_\s-]*MANUTEN", re.IGNORECASE),
        "FILIAL_PROJETO":     re.compile(r"FILIAL[_\s-]*PROJ",    re.IGNORECASE),
        "MATRIZ_MANUTENÇÃO":  re.compile(r"MATRIZ[_\s-]*MANUTEN", re.IGNORECASE),
        "MATRIZ_PROJETO":     re.compile(r"MATRIZ[_\s-]*PROJ",    re.IGNORECASE),
    }

    def detect_frente(filename: str) -> Optional[str]:
        for frente, pattern in frente_patterns.items():
            if pattern.search(filename):
                return frente
        return None

    def extract_provider(filename: str) -> str:
        """Extrai nome do prestador: tudo antes de _FILIAL_ ou _MATRIZ_."""
        base = os.path.splitext(filename)[0]
        # Encontra a posição onde começa FILIAL ou MATRIZ
        match = re.search(r'[_\s-]+(FILIAL|MATRIZ)[_\s-]', base, re.IGNORECASE)
        if match:
            provider = base[:match.start()].strip("_- ")
            return provider.upper() if provider else "DESCONHECIDO"
        return base.upper()

    def scan_dir(directory: str, version: str):
        if not os.path.isdir(directory):
            return
        for fname in os.listdir(directory):
            if not fname.lower().endswith((".xlsm", ".xlsx")):
                continue
            frente = detect_frente(fname)
            if not frente:
                continue
            provider = extract_provider(fname)
            if not provider or provider in ("DESCONHECIDO", ""):
                continue
            if provider not in result["files"]:
                result["files"][provider] = {"V1":{}, "V2":{}, "V3":{}}
            result["files"][provider][version][frente] = os.path.join(directory, fname)

    # V1 = raiz da pasta
    scan_dir(base_path, "V1")
    # V2 = subpasta V2
    scan_dir(os.path.join(base_path, "V2"), "V2")
    # V3 = subpasta V3/Projeto (ignora Geral)
    v3_proj = os.path.join(base_path, "V3", "Projeto")
    if os.path.isdir(v3_proj):
        scan_dir(v3_proj, "V3")
    else:
        # Fallback: tenta diretamente V3/ (ignorando subpasta Geral)
        v3_base = os.path.join(base_path, "V3")
        if os.path.isdir(v3_base):
            for item in os.listdir(v3_base):
                if item.lower() == "geral":
                    continue
                sub = os.path.join(v3_base, item)
                if os.path.isdir(sub):
                    scan_dir(sub, "V3")
                elif item.lower().endswith((".xlsm", ".xlsx")):
                    # Arquivo solto em V3
                    frente = detect_frente(item)
                    if frente:
                        prov = extract_provider(item, frente)
                        if prov:
                            result["files"].setdefault(prov, {"V1":{}, "V2":{}, "V3":{}})
                            result["files"][prov]["V3"][frente] = os.path.join(v3_base, item)

    result["providers"] = sorted(result["files"].keys())

    # Avisos de arquivos faltando
    for prov, versions in result["files"].items():
        for ver in ["V1", "V2", "V3"]:
            for frente in FRENTES:
                if frente not in versions[ver]:
                    result["warnings"].append(
                        f"⚠️ Arquivo não encontrado: {prov} / {ver} / {frente}"
                    )
    return result


def validate_provider(folder_path: str, provider: str, scan: dict) -> dict:
    """
    Valida todas as 4 frentes × 3 versões para um prestador.
    Retorna dict completo com resultados por frente/versão.
    """
    files = scan["files"].get(provider, {})
    resultado = {
        "provider": provider,
        "folder":   folder_path,
        "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "frentes":  {},
        "ok_global": True,
        "total_erros": 0,
    }

    for frente in FRENTES:
        resultado["frentes"][frente] = {
            "V1": None, "V2": None, "V3": None,
            "ok": True, "erros_total": 0
        }
        for ver in ["V1", "V2", "V3"]:
            filepath = files.get(ver, {}).get(frente)
            if not filepath or not os.path.exists(filepath):
                res = {
                    "ok": False,
                    "total_geral": 0,
                    "erros": [{
                        "linha": "-", "celula": "-",
                        "tipo": "📁 Arquivo não encontrado",
                        "encontrado": "Arquivo ausente",
                        "esperado":   f"Arquivo {ver} para {frente}",
                        "como_corrigir":
                            f"Adicione o arquivo de {ver} para a frente {frente} "
                            f"do prestador {provider} na pasta correta:\n"
                            f"• V1: raiz da pasta do mês\n"
                            f"• V2: subpasta V2/\n"
                            f"• V3: subpasta V3/Projeto/"
                    }]
                }
            else:
                res = validate_single_file(filepath)
                res["filepath"] = filepath

            resultado["frentes"][frente][ver] = res
            if not res["ok"]:
                resultado["frentes"][frente]["ok"] = False
                resultado["frentes"][frente]["erros_total"] += len(res["erros"])
                resultado["ok_global"] = False
                resultado["total_erros"] += len(res["erros"])

    return resultado


# ══════════════════════════════════════════════════════════════
#  INTERFACE GRÁFICA
# ══════════════════════════════════════════════════════════════

class ScrollableFrame(tk.Frame):
    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self.canvas  = tk.Canvas(self, bg=C_PANEL, highlightthickness=0)
        self.vscroll = ttk.Scrollbar(self, orient="vertical",
                                     command=self.canvas.yview)
        self.inner   = tk.Frame(self.canvas, bg=C_PANEL)
        self.inner.bind("<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0,0), window=self.inner, anchor="nw")
        self.canvas.configure(yscrollcommand=self.vscroll.set)
        self.canvas.pack(side="left",  fill="both", expand=True)
        self.vscroll.pack(side="right", fill="y")
        self.canvas.bind("<MouseWheel>",
            lambda e: self.canvas.yview_scroll(-1*(e.delta//120), "units"))


class StatusDot(tk.Label):
    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self.set_idle()

    def set_idle(self):
        self.config(text="⬤", fg=C_MUTED)

    def set_ok(self):
        self.config(text="⬤", fg=C_GREEN)

    def set_fail(self):
        self.config(text="⬤", fg=C_RED)

    def set_warn(self):
        self.config(text="⬤", fg=C_ORANGE)


class ValidadorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Validador de Planilhas — Prestadores")
        self.geometry("1100x750")
        self.minsize(900, 600)
        self.configure(bg=C_BG)

        self._scan_data       = None
        self._result          = None
        self._ref_comparacoes = {}
        self._folder_path     = tk.StringVar()
        self._provider        = tk.StringVar()
        self._ref_path        = tk.StringVar()
        self._status_msg      = tk.StringVar(value="Selecione a pasta do mês para começar")

        self._build_styles()
        self._build_ui()

    # ─── Estilos ───────────────────────────────────────────────────
    def _build_styles(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure(".",         background=C_BG,    foreground=C_TEXT,
                     font=("Segoe UI", 10))
        s.configure("TFrame",    background=C_BG)
        s.configure("TLabel",    background=C_BG,    foreground=C_TEXT)
        s.configure("TButton",   background=C_ACCENT, foreground=C_BG,
                     font=("Segoe UI", 10, "bold"), padding=(12,6))
        s.map("TButton",
              background=[("active", "#9B9FFF"), ("disabled", C_BORDER)],
              foreground=[("disabled", C_MUTED)])
        s.configure("Accent.TButton", background=C_ACCENT, foreground=C_BG)
        s.configure("TCombobox", fieldbackground=C_CARD, background=C_CARD,
                     foreground=C_TEXT, selectbackground=C_ACCENT)
        s.configure("Treeview", background=C_CARD, foreground=C_TEXT,
                     fieldbackground=C_CARD, rowheight=26,
                     font=("Segoe UI", 9))
        s.configure("Treeview.Heading", background=C_PANEL, foreground=C_ACCENT,
                     font=("Segoe UI", 9, "bold"))
        s.map("Treeview", background=[("selected", C_ACCENT)],
              foreground=[("selected", C_BG)])
        s.configure("TScrollbar", background=C_BORDER, troughcolor=C_PANEL)
        s.configure("TEntry", fieldbackground=C_CARD, foreground=C_TEXT,
                     insertcolor=C_TEXT)

    # ─── UI principal ──────────────────────────────────────────────
    def _build_ui(self):
        # Título
        top = tk.Frame(self, bg=C_PANEL, pady=0)
        top.pack(fill="x")
        tk.Label(top, text="  ✦ Validador de Planilhas",
                 font=("Segoe UI", 15, "bold"),
                 bg=C_PANEL, fg=C_ACCENT, pady=10).pack(side="left")
        tk.Label(top, text="Prestadores · 4 Frentes · V1 / V2 / V3",
                 font=("Segoe UI", 9), bg=C_PANEL, fg=C_MUTED, pady=10).pack(side="left", padx=8)

        sep = tk.Frame(self, bg=C_BORDER, height=1)
        sep.pack(fill="x")

        # Corpo = painel esquerdo + direito
        body = tk.Frame(self, bg=C_BG)
        body.pack(fill="both", expand=True, padx=0, pady=0)
        body.columnconfigure(0, minsize=320, weight=0)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        # Painel esquerdo (controles)
        left = tk.Frame(body, bg=C_PANEL, width=320)
        left.grid(row=0, column=0, sticky="nsew")
        left.pack_propagate(False)
        self._build_left(left)

        # Separador vertical
        tk.Frame(body, bg=C_BORDER, width=1).grid(row=0, column=0, sticky="nse")

        # Painel direito (resultados)
        right = tk.Frame(body, bg=C_BG)
        right.grid(row=0, column=1, sticky="nsew")
        self._build_right(right)

        # Status bar
        bar = tk.Frame(self, bg=C_PANEL, pady=4)
        bar.pack(fill="x", side="bottom")
        tk.Frame(bar, bg=C_BORDER, height=1).pack(fill="x")
        tk.Label(bar, textvariable=self._status_msg,
                 font=("Segoe UI", 9), bg=C_PANEL, fg=C_MUTED, padx=12).pack(side="left")
        tk.Label(bar, text="Validador de Planilhas v2.0",
                 font=("Segoe UI", 9), bg=C_PANEL, fg=C_BORDER, padx=12).pack(side="right")

    # ─── Painel esquerdo ───────────────────────────────────────────
    def _build_left(self, parent):
        pad = {"padx": 16, "pady": 6}

        # Seção pasta
        tk.Label(parent, text="📁  PASTA DO MÊS",
                 font=("Segoe UI", 9, "bold"),
                 bg=C_PANEL, fg=C_ACCENT, **pad).pack(anchor="w", pady=(16,4))

        frm_path = tk.Frame(parent, bg=C_PANEL)
        frm_path.pack(fill="x", **pad)
        entry = tk.Entry(frm_path, textvariable=self._folder_path,
                         font=("Segoe UI", 9), bg=C_CARD, fg=C_TEXT,
                         insertbackground=C_TEXT, relief="flat", bd=4)
        entry.pack(side="left", fill="x", expand=True, ipady=4)
        tk.Button(frm_path, text="...",
                  font=("Segoe UI", 9, "bold"),
                  bg=C_ACCENT, fg=C_BG, relief="flat", bd=0,
                  padx=8, pady=4,
                  command=self._select_folder).pack(side="left", padx=(4,0))

        self._btn_scan = tk.Button(parent, text="⟳  Carregar Pasta",
                                   font=("Segoe UI", 10, "bold"),
                                   bg=C_CARD, fg=C_ACCENT,
                                   relief="flat", bd=0, pady=8,
                                   command=self._load_folder)
        self._btn_scan.pack(fill="x", **pad)

        # Info da pasta
        self._lbl_folder_info = tk.Label(parent, text="",
                                          font=("Segoe UI", 8),
                                          bg=C_PANEL, fg=C_MUTED,
                                          wraplength=280, justify="left")
        self._lbl_folder_info.pack(anchor="w", padx=16)

        # Separador
        tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=16, pady=12)

        # Seção prestador
        tk.Label(parent, text="👷  PRESTADOR",
                 font=("Segoe UI", 9, "bold"),
                 bg=C_PANEL, fg=C_ACCENT, **pad).pack(anchor="w", pady=(0,4))

        self._cmb_provider = ttk.Combobox(parent, textvariable=self._provider,
                                           state="disabled",
                                           font=("Segoe UI", 11, "bold"))
        self._cmb_provider.pack(fill="x", **pad)

        # Avisos de arquivos
        self._lbl_warnings = tk.Label(parent, text="",
                                       font=("Segoe UI", 8),
                                       bg=C_PANEL, fg=C_ORANGE,
                                       wraplength=280, justify="left")
        self._lbl_warnings.pack(anchor="w", padx=16)

        tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=16, pady=12)

        # Seção planilha de referência
        tk.Label(parent, text="📊  PLANILHA DE REFERÊNCIA",
                 font=("Segoe UI", 9, "bold"),
                 bg=C_PANEL, fg=C_ACCENT, **pad).pack(anchor="w", pady=(0,4))

        frm_ref = tk.Frame(parent, bg=C_PANEL)
        frm_ref.pack(fill="x", **pad)
        entry_ref = tk.Entry(frm_ref, textvariable=self._ref_path,
                             font=("Segoe UI", 8), bg=C_CARD, fg=C_TEXT,
                             insertbackground=C_TEXT, relief="flat", bd=4)
        entry_ref.pack(side="left", fill="x", expand=True, ipady=3)
        tk.Button(frm_ref, text="...",
                  font=("Segoe UI", 9, "bold"),
                  bg=C_ACCENT, fg=C_BG, relief="flat", bd=0,
                  padx=8, pady=3,
                  command=self._select_ref_file).pack(side="left", padx=(4,0))

        self._lbl_ref_info = tk.Label(parent, text="Opcional — usado para comparar totais" if _REF_DISPONIVEL
                                      else "⚠️ Módulo referencia.py não encontrado",
                                      font=("Segoe UI", 8),
                                      bg=C_PANEL, fg=C_MUTED if _REF_DISPONIVEL else C_ORANGE,
                                      wraplength=280, justify="left")
        self._lbl_ref_info.pack(anchor="w", padx=16)

        tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=16, pady=12)

        # Botão validar
        self._btn_validate = tk.Button(parent, text="▶  VALIDAR",
                                        font=("Segoe UI", 13, "bold"),
                                        bg=C_GREEN, fg=C_BG,
                                        relief="flat", bd=0, pady=10,
                                        state="disabled",
                                        command=self._run_validation)
        self._btn_validate.pack(fill="x", padx=16, pady=4)

        # Progress
        self._progress = ttk.Progressbar(parent, mode="indeterminate")
        self._progress.pack(fill="x", padx=16, pady=4)

        tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=16, pady=12)

        # Botão exportar
        self._btn_export = tk.Button(parent, text="💾  Exportar Relatório",
                                      font=("Segoe UI", 10),
                                      bg=C_CARD, fg=C_MUTED,
                                      relief="flat", bd=0, pady=7,
                                      state="disabled",
                                      command=self._export_report)
        self._btn_export.pack(fill="x", padx=16)

        # Resumo rápido
        tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=16, pady=12)
        self._frm_summary_quick = tk.Frame(parent, bg=C_PANEL)
        self._frm_summary_quick.pack(fill="x", padx=16)

    # ─── Painel direito ────────────────────────────────────────────
    def _build_right(self, parent):
        parent.rowconfigure(1, weight=1)
        parent.columnconfigure(0, weight=1)

        # Banner de resultado global
        self._frm_banner = tk.Frame(parent, bg=C_CARD, pady=12)
        self._frm_banner.grid(row=0, column=0, sticky="ew", padx=0, pady=0)
        self._lbl_banner = tk.Label(self._frm_banner,
                                     text="Aguardando validação...",
                                     font=("Segoe UI", 14, "bold"),
                                     bg=C_CARD, fg=C_MUTED)
        self._lbl_banner.pack()
        self._lbl_banner_sub = tk.Label(self._frm_banner, text="",
                                         font=("Segoe UI", 9),
                                         bg=C_CARD, fg=C_MUTED)
        self._lbl_banner_sub.pack()

        # Notebook = abas por frente
        self._notebook = ttk.Notebook(parent)
        self._notebook.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)

        # Criar aba inicial
        self._tab_frames = {}
        for frente in FRENTES:
            frame = tk.Frame(self._notebook, bg=C_BG)
            self._notebook.add(frame, text=f"  {FRENTE_LABELS[frente]}  ")
            self._tab_frames[frente] = frame
            self._build_frente_tab(frame, frente, empty=True)

        style = ttk.Style()
        style.configure("TNotebook",        background=C_PANEL, borderwidth=0)
        style.configure("TNotebook.Tab",    background=C_CARD,  foreground=C_MUTED,
                         padding=(14,8), font=("Segoe UI", 9))
        style.map("TNotebook.Tab",
                  background=[("selected", C_BG)],
                  foreground=[("selected", C_ACCENT)])

    def _build_frente_tab(self, parent, frente, empty=False):
        """Constrói (ou reconstrói) o conteúdo de uma aba de frente."""
        for w in parent.winfo_children():
            w.destroy()

        if empty:
            tk.Label(parent, text="Nenhuma validação realizada ainda.",
                     font=("Segoe UI", 10), bg=C_BG, fg=C_MUTED).pack(pady=40)
            return

        result = self._result
        if not result:
            return

        frente_data = result["frentes"].get(frente, {})
        frente_ok   = frente_data.get("ok", True)

        # Cabeçalho da frente
        hdr = tk.Frame(parent, bg=C_PANEL, pady=8)
        hdr.pack(fill="x")
        dot = "✅" if frente_ok else "❌"
        status_txt = "APROVADO" if frente_ok else f"REPROVADO — {frente_data.get('erros_total',0)} erro(s)"
        tk.Label(hdr, text=f"  {dot}  {FRENTE_LABELS[frente]}",
                 font=("Segoe UI", 12, "bold"), bg=C_PANEL,
                 fg=C_GREEN if frente_ok else C_RED).pack(side="left", padx=8)
        tk.Label(hdr, text=status_txt, font=("Segoe UI", 10),
                 bg=C_PANEL, fg=C_GREEN if frente_ok else C_RED).pack(side="left")

        # Cards de versão (V1, V2, V3)
        cards_frm = tk.Frame(parent, bg=C_BG)
        cards_frm.pack(fill="x", padx=12, pady=8)

        # Comparações de referência para esta frente (se disponíveis)
        ref_frente = self._ref_comparacoes.get(frente, {})
        ref_comps_by_ver = {c["versao"]: c
                            for c in ref_frente.get("comparacoes", [])}

        for i, ver in enumerate(["V1","V2","V3"]):
            ver_data = frente_data.get(ver)
            ver_ok   = ver_data.get("ok", False) if ver_data else False
            n_erros  = len(ver_data.get("erros", [])) if ver_data else 0
            tg       = ver_data.get("total_geral", 0) if ver_data else 0

            # Cor de borda: considera também comparação de referência
            ref_c    = ref_comps_by_ver.get(ver, {})
            ref_match = ref_c.get("match")   # True/False/None
            if ref_match is False:
                border_col = C_ORANGE     # planilha ok mas diverge da referência
            elif ver_ok:
                border_col = C_GREEN
            else:
                border_col = C_RED

            card = tk.Frame(cards_frm, bg=C_CARD, padx=12, pady=8,
                            relief="flat", bd=0,
                            highlightthickness=2,
                            highlightbackground=border_col)
            card.grid(row=0, column=i, padx=6, sticky="ew")
            cards_frm.columnconfigure(i, weight=1)

            tk.Label(card, text=VERSION_LABELS[ver],
                     font=("Segoe UI", 9, "bold"), bg=C_CARD,
                     fg=C_ACCENT).pack(anchor="w")

            status_icon = "✅ OK" if ver_ok else f"❌ {n_erros} erro(s)"
            tk.Label(card, text=status_icon,
                     font=("Segoe UI", 11, "bold"), bg=C_CARD,
                     fg=C_GREEN if ver_ok else C_RED).pack(anchor="w", pady=4)

            if tg:
                tk.Label(card, text=f"Total Geral: R${tg:,.2f}",
                         font=("Segoe UI", 9), bg=C_CARD, fg=C_MUTED).pack(anchor="w")

            # Referência
            if ref_c:
                ref_val = ref_c.get("referencia")
                ref_err = ref_c.get("erro")
                if ref_err:
                    tk.Label(card, text=f"Ref: {ref_err[:45]}",
                             font=("Segoe UI", 7), bg=C_CARD,
                             fg=C_MUTED, wraplength=180).pack(anchor="w", pady=(2,0))
                elif ref_val is not None:
                    dif = ref_c.get("diferenca", 0)
                    if ref_match:
                        ref_txt = f"Referência: R${ref_val:,.2f} ✅"
                        ref_fg  = C_GREEN
                    else:
                        ref_txt = f"Referência: R${ref_val:,.2f}  ⚠️ Δ R${dif:,.2f}"
                        ref_fg  = C_ORANGE
                    tk.Label(card, text=ref_txt,
                             font=("Segoe UI", 9, "bold"), bg=C_CARD,
                             fg=ref_fg).pack(anchor="w", pady=(2,0))

            if ver_data and ver_data.get("filepath"):
                fname = os.path.basename(ver_data["filepath"])
                tk.Label(card, text=fname, font=("Segoe UI", 7),
                         bg=C_CARD, fg=C_BORDER, wraplength=180).pack(anchor="w", pady=(4,0))

        # Tabela de erros
        if frente_data.get("erros_total", 0) > 0:
            tk.Label(parent, text="  DETALHES DOS ERROS",
                     font=("Segoe UI", 9, "bold"), bg=C_BG, fg=C_ORANGE).pack(
                         anchor="w", padx=12, pady=(8,2))

            tree_frm = tk.Frame(parent, bg=C_BG)
            tree_frm.pack(fill="both", expand=True, padx=12, pady=(0,8))

            cols_tree = ("ver","linha","celula","tipo","encontrado","esperado")
            tree = ttk.Treeview(tree_frm, columns=cols_tree, show="headings",
                                selectmode="browse")
            tree.heading("ver",        text="Versão")
            tree.heading("linha",      text="Linha")
            tree.heading("celula",     text="Célula")
            tree.heading("tipo",       text="Tipo de Erro")
            tree.heading("encontrado", text="Encontrado")
            tree.heading("esperado",   text="Esperado")

            tree.column("ver",        width=90,  anchor="center")
            tree.column("linha",      width=60,  anchor="center")
            tree.column("celula",     width=70,  anchor="center")
            tree.column("tipo",       width=220, anchor="w")
            tree.column("encontrado", width=130, anchor="center")
            tree.column("esperado",   width=180, anchor="w")

            vscr = ttk.Scrollbar(tree_frm, orient="vertical", command=tree.yview)
            hscr = ttk.Scrollbar(tree_frm, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vscr.set, xscrollcommand=hscr.set)

            tree.grid(row=0, column=0, sticky="nsew")
            vscr.grid(row=0, column=1, sticky="ns")
            hscr.grid(row=1, column=0, sticky="ew")
            tree_frm.rowconfigure(0, weight=1)
            tree_frm.columnconfigure(0, weight=1)

            tree.tag_configure("erro",  background="#3A1A1A", foreground=C_RED)
            tree.tag_configure("warn",  background="#3A2A00", foreground=C_ORANGE)
            tree.tag_configure("hard",  background="#2A1A00", foreground=C_YELLOW)

            # Preencher árvore
            for ver in ["V1","V2","V3"]:
                ver_data = frente_data.get(ver, {})
                if not ver_data:
                    continue
                for err in ver_data.get("erros", []):
                    tipo = err["tipo"]
                    if "fórmula" in tipo.lower() or "excel" in tipo.lower():
                        tag = "erro"
                    elif "hardcoded" in tipo.lower():
                        tag = "hard"
                    else:
                        tag = "warn"
                    tree.insert("", "end", values=(
                        VERSION_LABELS[ver],
                        err["linha"], err["celula"],
                        tipo,
                        err["encontrado"], err["esperado"]
                    ), tags=(tag,))

            # Painel "Como corrigir" ao selecionar linha
            lbl_fix_title = tk.Label(parent, text="  🔧  COMO CORRIGIR",
                                      font=("Segoe UI", 9, "bold"),
                                      bg=C_BG, fg=C_ACCENT)
            lbl_fix_title.pack(anchor="w", padx=12, pady=(4,0))
            lbl_fix = tk.Label(parent, text="Selecione um erro acima para ver como corrigir.",
                                font=("Segoe UI", 9), bg=C_CARD, fg=C_MUTED,
                                wraplength=700, justify="left", padx=12, pady=8,
                                anchor="w")
            lbl_fix.pack(fill="x", padx=12, pady=(0,8))

            # Montar index de correções
            fix_map = {}
            idx = 0
            for ver in ["V1","V2","V3"]:
                vd = frente_data.get(ver, {})
                for err in (vd.get("erros", []) if vd else []):
                    fix_map[idx] = err.get("como_corrigir","")
                    idx += 1

            def on_select(event, t=tree, lbl=lbl_fix, fm=fix_map):
                sel = t.selection()
                if sel:
                    row_idx_sel = t.index(sel[0])
                    fix = fm.get(row_idx_sel, "")
                    lbl.config(text=fix or "Sem detalhes disponíveis.",
                               fg=C_YELLOW)

            tree.bind("<<TreeviewSelect>>", on_select)
        else:
            tk.Label(parent, text="✅  Nenhum erro encontrado nesta frente.",
                     font=("Segoe UI", 11), bg=C_BG, fg=C_GREEN).pack(pady=30)

    # ─── Ações ─────────────────────────────────────────────────────
    def _select_folder(self):
        path = filedialog.askdirectory(title="Selecione a pasta do mês")
        if path:
            self._folder_path.set(path)
            self._load_folder()

    def _select_ref_file(self):
        path = filedialog.askopenfilename(
            title="Selecione a planilha de referência (IMPOSTO)",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")]
        )
        if path:
            self._ref_path.set(path)
            fname = os.path.basename(path)
            self._lbl_ref_info.config(text=f"✅ {fname}", fg=C_GREEN)

    def _load_folder(self):
        path = self._folder_path.get().strip()
        if not path or not os.path.isdir(path):
            messagebox.showwarning("Pasta inválida",
                "Selecione uma pasta válida.")
            return

        self._set_status("Carregando pasta...")
        self._progress.start(10)
        self._btn_scan.config(state="disabled")

        def do_scan():
            scan = scan_folder(path)
            self.after(0, lambda: self._on_scan_done(scan, path))

        threading.Thread(target=do_scan, daemon=True).start()

    def _on_scan_done(self, scan, path):
        self._progress.stop()
        self._btn_scan.config(state="normal")
        self._scan_data = scan

        providers = scan["providers"]
        if not providers:
            self._set_status("⚠️ Nenhum arquivo de planilha encontrado nesta pasta.")
            self._lbl_folder_info.config(
                text="Nenhum prestador encontrado. Verifique a estrutura da pasta.")
            self._cmb_provider.config(state="disabled")
            self._btn_validate.config(state="disabled")
            return

        folder_name = os.path.basename(path)
        n_files     = sum(
            len(v) for pdata in scan["files"].values()
            for v in pdata.values()
        )
        self._lbl_folder_info.config(
            text=f"📂 {folder_name}\n"
                 f"👷 {len(providers)} prestador(es) · 📄 {n_files} arquivo(s)"
        )

        self._cmb_provider["values"] = providers
        self._cmb_provider.config(state="readonly")
        if len(providers) == 1:
            self._provider.set(providers[0])
            self._btn_validate.config(state="normal")
        else:
            self._provider.set("")
            self._cmb_provider.bind("<<ComboboxSelected>>",
                lambda e: self._btn_validate.config(state="normal"))

        # Mostrar avisos
        warns = scan.get("warnings", [])
        if warns:
            self._lbl_warnings.config(
                text="\n".join(warns[:5]) + ("\n..." if len(warns) > 5 else ""))
        else:
            self._lbl_warnings.config(text="")

        self._set_status(f"Pasta carregada: {len(providers)} prestador(es) encontrado(s).")

    def _run_validation(self):
        provider = self._provider.get().strip()
        path     = self._folder_path.get().strip()
        if not provider or not self._scan_data:
            messagebox.showwarning("Atenção", "Selecione um prestador.")
            return

        self._set_status(f"Validando {provider}...")
        self._btn_validate.config(state="disabled")
        self._btn_export.config(state="disabled")
        self._progress.start(10)

        ref_path     = self._ref_path.get().strip()
        folder_name  = os.path.basename(path)

        def do_validate():
            result = validate_provider(path, provider, self._scan_data)
            ref_comps = {}
            if _REF_DISPONIVEL and ref_path and os.path.isfile(ref_path):
                try:
                    ref_comps = _comparar_ref(ref_path, folder_name, provider, result)
                except Exception as exc:
                    ref_comps = {"_erro": str(exc)}
            self.after(0, lambda: self._on_validation_done(result, ref_comps))

        threading.Thread(target=do_validate, daemon=True).start()

    def _on_validation_done(self, result, ref_comps=None):
        self._progress.stop()
        self._btn_validate.config(state="normal")
        self._btn_export.config(state="normal")
        self._result = result
        self._ref_comparacoes = ref_comps or {}

        ok      = result["ok_global"]
        n_erros = result["total_erros"]
        prov    = result["provider"]

        # Banner
        # Verificar se alguma referência diverge
        ref_divergencias = sum(
            1 for fd in self._ref_comparacoes.values()
            if fd.get("comparacoes")
            for c in fd["comparacoes"]
            if c.get("match") is False
        )

        if ok and ref_divergencias == 0:
            self._frm_banner.config(bg="#0A2A0A")
            self._lbl_banner.config(
                text=f"✅  {prov}  —  APROVADO",
                font=("Segoe UI", 15, "bold"),
                bg="#0A2A0A", fg=C_GREEN)
            sub_txt = "Todas as 4 frentes passaram na validação."
            if self._ref_comparacoes and "_erro" not in self._ref_comparacoes:
                sub_txt += "  ·  Totais conferem com a referência."
            self._lbl_banner_sub.config(text=sub_txt, bg="#0A2A0A", fg=C_MUTED)
        elif ok and ref_divergencias > 0:
            self._frm_banner.config(bg="#2A1A00")
            self._lbl_banner.config(
                text=f"⚠️  {prov}  —  APROVADO COM DIVERGÊNCIAS",
                font=("Segoe UI", 14, "bold"),
                bg="#2A1A00", fg=C_ORANGE)
            self._lbl_banner_sub.config(
                text=f"Fórmulas OK mas {ref_divergencias} total(is) divergem da planilha de referência.",
                bg="#2A1A00", fg=C_ORANGE)
        else:
            self._frm_banner.config(bg="#2A0A0A")
            self._lbl_banner.config(
                text=f"❌  {prov}  —  REPROVADO",
                font=("Segoe UI", 15, "bold"),
                bg="#2A0A0A", fg=C_RED)
            sub_txt = f"{n_erros} erro(s) encontrado(s). Veja os detalhes nas abas abaixo."
            if ref_divergencias > 0:
                sub_txt += f"  ·  {ref_divergencias} divergência(s) com a referência."
            self._lbl_banner_sub.config(text=sub_txt, bg="#2A0A0A", fg=C_ORANGE)

        # Atualizar abas
        for frente in FRENTES:
            frame = self._tab_frames[frente]
            self._build_frente_tab(frame, frente, empty=False)

            # Atualizar título da aba
            fd_ok = result["frentes"].get(frente, {}).get("ok", True)
            icon  = "✅" if fd_ok else "❌"
            idx   = FRENTES.index(frente)
            short = frente.replace("FILIAL_","F·").replace("MATRIZ_","M·")
            self._notebook.tab(idx, text=f"  {icon} {short}  ")

        # Resumo rápido no painel esquerdo
        for w in self._frm_summary_quick.winfo_children():
            w.destroy()
        tk.Label(self._frm_summary_quick, text="RESUMO",
                 font=("Segoe UI", 8, "bold"),
                 bg=C_PANEL, fg=C_ACCENT).pack(anchor="w")
        for frente in FRENTES:
            fd = result["frentes"].get(frente, {})
            fd_ok  = fd.get("ok", True)
            n_err  = fd.get("erros_total", 0)
            icon   = "✅" if fd_ok else "❌"
            color  = C_GREEN if fd_ok else C_RED
            short  = frente.replace("FILIAL_", "Filial·").replace("MATRIZ_", "Matriz·")
            txt    = f"{icon} {short}" if fd_ok else f"{icon} {short} ({n_err})"
            tk.Label(self._frm_summary_quick, text=txt,
                     font=("Segoe UI", 9), bg=C_PANEL, fg=color).pack(anchor="w")

        self._set_status(
            f"Validação concluída: {prov} — {'APROVADO ✅' if ok else f'REPROVADO ❌ ({n_erros} erros)'}"
            f"  ·  {result['timestamp']}"
        )

    # ─── Exportar Relatório ────────────────────────────────────────
    def _export_report(self):
        if not self._result:
            return
        outpath = filedialog.asksaveasfilename(
            title="Salvar relatório",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"RELATORIO_{self._result['provider']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        if not outpath:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "RELATÓRIO"

            r = self._result
            ok = r["ok_global"]

            # Título
            ws.merge_cells("A1:G1")
            ws["A1"] = f"VALIDAÇÃO — {r['provider']}  —  {'✅ APROVADO' if ok else '❌ REPROVADO'}"
            ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
            ws["A1"].fill = PatternFill("solid", start_color="1F3864" if ok else "C00000")
            ws["A1"].alignment = Alignment(horizontal="center")

            ws["A2"] = f"Pasta: {r['folder']}  ·  Gerado em: {r['timestamp']}"
            ws["A2"].font = Font(name="Arial", size=9, italic=True, color="888888")
            ws.merge_cells("A2:G2")

            # Cabeçalho
            hdrs = ["FRENTE", "VERSÃO", "LINHA", "CÉLULA",
                    "TIPO DE ERRO", "ENCONTRADO", "ESPERADO", "COMO CORRIGIR",
                    "TOTAL GERAL", "REFERÊNCIA", "DIFERENÇA", "CONF. REF."]
            for i, h in enumerate(hdrs, 1):
                c = ws.cell(4, i, h)
                c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
                c.fill = PatternFill("solid", start_color="2E75B6")
                c.alignment = Alignment(horizontal="center")

            ref_comps = self._ref_comparacoes

            row = 5
            for frente in FRENTES:
                fd = r["frentes"].get(frente, {})
                ref_fd = ref_comps.get(frente, {})
                ref_by_ver = {c["versao"]: c for c in ref_fd.get("comparacoes", [])}

                for ver in ["V1","V2","V3"]:
                    vd = fd.get(ver) or {}
                    rc = ref_by_ver.get(ver, {})
                    tg_str   = f"R${vd.get('total_geral',0):,.2f}" if vd else ""
                    ref_val  = rc.get("referencia")
                    ref_str  = f"R${ref_val:,.2f}" if ref_val is not None else (rc.get("erro","") or "")
                    dif_str  = f"R${rc.get('diferenca',0):,.2f}" if rc.get("diferenca") is not None else ""
                    match_str= ("✅" if rc.get("match") else ("❌" if rc.get("match") is False else ""))

                    for err in vd.get("erros", []):
                        bg = "FFCCCC" if "erro" in err["tipo"].lower() else \
                             "FFF2CC" if "hardcoded" in err["tipo"].lower() else "FFE5CC"
                        vals = [FRENTE_LABELS[frente], VERSION_LABELS[ver],
                                err["linha"], err["celula"], err["tipo"],
                                str(err["encontrado"]), str(err["esperado"]),
                                err.get("como_corrigir",""),
                                tg_str, ref_str, dif_str, match_str]
                        for i, v in enumerate(vals, 1):
                            c = ws.cell(row, i, v)
                            c.font = Font(name="Arial", size=9)
                            c.fill = PatternFill("solid", start_color=bg)
                            c.alignment = Alignment(wrap_text=True)
                        row += 1

                    if not vd.get("erros"):
                        ver_ok_str = "✅ Aprovado"
                        bg = "E2EFDA" if rc.get("match") is not False else "FFF2CC"
                        vals = [FRENTE_LABELS[frente], VERSION_LABELS[ver],
                                "", "", ver_ok_str, "", "", "",
                                tg_str, ref_str, dif_str, match_str]
                        for i, v in enumerate(vals, 1):
                            c = ws.cell(row, i, v)
                            c.fill = PatternFill("solid", start_color=bg)
                            c.font = Font(name="Arial", size=9,
                                          color="375623" if bg == "E2EFDA" else "7F4E00")
                        row += 1

            for col_idx, width in enumerate([30, 14, 8, 8, 28, 18, 22, 50, 14, 14, 12, 10], 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            wb.save(outpath)
            messagebox.showinfo("Exportado!",
                f"Relatório salvo em:\n{outpath}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e))

    def _set_status(self, msg: str):
        self._status_msg.set(msg)
        self.update_idletasks()


# ══════════════════════════════════════════════════════════════
#  PONTO DE ENTRADA
# ══════════════════════════════════════════════════════════════

def main():
    app = ValidadorApp()
    app.mainloop()

if __name__ == "__main__":
    main()
