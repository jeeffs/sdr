"""
Módulo de leitura da planilha de referência IMPOSTO.

Estrutura da aba de mês:
  Linha 1: Cabeçalhos de seção (VALORES MEGALINK=V3, VALORES TERCEIROS=V1)
  Linha 4: MATRIZ + PRESTADOR + PROJETO + MANUTENÇÃO (para V3 e V1)
  Linhas 5...: Dados por prestador (MATRIZ)
  Linha 14ish: FILIAL seção
  Linhas 15...: Dados por prestador (FILIAL)
  Linha 28ish: Seção V2 (abaixo de V1), colunas F=PROJETO, G=MANU, col I=PRESTADOR
"""

import openpyxl
import re
from typing import Optional

# Meses em português para mapear pasta→aba
MESES_PT = {
    "JAN": "JANEIRO", "FEV": "FEVEREIRO", "MAR": "MARÇO",
    "ABR": "ABRIL",   "MAI": "MAIO",      "JUN": "JUNHO",
    "JUL": "JULHO",   "AGO": "AGOSTO",    "SET": "SETEMBRO",
    "OUT": "OUTUBRO", "NOV": "NOVEMBRO",  "DEZ": "DEZEMBRO",
    "01": "JANEIRO",  "02": "FEVEREIRO",  "03": "MARÇO",
    "04": "ABRIL",    "05": "MAIO",       "06": "JUNHO",
    "07": "JULHO",    "08": "AGOSTO",     "09": "SETEMBRO",
    "10": "OUTUBRO",  "11": "NOVEMBRO",   "12": "DEZEMBRO",
}

def normalizar(texto) -> str:
    if texto is None:
        return ""
    return " ".join(str(texto).strip().split()).upper()


def encontrar_aba_mes(wb, folder_name: str) -> Optional[str]:
    """
    Tenta mapear o nome da pasta (ex: 'FEV_2026', '2026-02', 'FEVEREIRO 2026')
    para o nome de uma aba no workbook.
    """
    nome = folder_name.upper().replace("-", "_").replace("/", "_")

    # Extrair mês e ano da pasta
    mes_encontrado = None
    ano_encontrado = None

    # Tentar padrão ANO_MES ou MES_ANO
    partes = re.split(r"[_\s\-/]", nome)
    for p in partes:
        if len(p) == 4 and p.isdigit():
            ano_encontrado = p[-2:]  # Últimos 2 dígitos
        elif p in MESES_PT:
            mes_encontrado = MESES_PT[p]
        elif p.isdigit() and 1 <= int(p) <= 12:
            mes_encontrado = MESES_PT[p.zfill(2)]
        elif len(p) >= 3:
            for abbr, full in MESES_PT.items():
                if p.startswith(abbr) or full.startswith(p[:3]):
                    mes_encontrado = full
                    break

    if not mes_encontrado or not ano_encontrado:
        return None

    # Candidatos de nome de aba
    candidatos = [
        f"{mes_encontrado} {ano_encontrado}",
        f"{mes_encontrado[:3]} {ano_encontrado}",
        f"{mes_encontrado} 20{ano_encontrado}",
    ]

    for tab in wb.sheetnames:
        tab_upper = tab.upper().strip()
        for c in candidatos:
            if c in tab_upper or tab_upper == c:
                return tab

    return None


def _detectar_colunas(ws) -> dict:
    """
    Detecta as colunas de V3 (MEGALINK), V1 (TERCEIROS) e V2
    lendo a linha 1 e linha 4.
    Retorna dict com posições 1-based.
    """
    cols = {
        "v3_proj": None,  "v3_manu": None,
        "v1_proj": None,  "v1_manu": None,
        "prestador_principal": None,
    }

    # Linha 1: encontrar VALORES MEGALINK e VALORES TERCEIROS
    row1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_megalink   = None
    col_terceiros  = None
    for i, v in enumerate(row1):
        s = str(v).upper() if v else ""
        if "MEGALINK" in s and col_megalink is None:
            col_megalink  = i + 1
        if "VALORES TERCEIROS" in s and col_terceiros is None:
            col_terceiros = i + 1

    # Linha 4: encontrar PROJETO e MANUTENÇÃO nas regiões de MEGALINK e TERCEIROS
    row4 = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
    for i, v in enumerate(row4):
        vnorm = normalizar(v)
        c1 = i + 1
        if vnorm in ("PRESTADOR",) and cols["prestador_principal"] is None:
            cols["prestador_principal"] = c1
        if col_megalink and c1 >= col_megalink and c1 < col_megalink + 5:
            if vnorm == "PROJETO" and cols["v3_proj"] is None:
                cols["v3_proj"] = c1
            if vnorm == "MANUTENÇÃO" and cols["v3_manu"] is None:
                cols["v3_manu"] = c1
        if col_terceiros and c1 >= col_terceiros and c1 < col_terceiros + 8:
            if vnorm == "PROJETO" and cols["v1_proj"] is None:
                cols["v1_proj"] = c1
            if vnorm == "MANUTENÇÃO" and cols["v1_manu"] is None:
                cols["v1_manu"] = c1

    return cols


def _encontrar_secao(ws, label: str, max_row: int = 120) -> Optional[int]:
    """Encontra a linha que marca início de uma seção (MATRIZ / FILIAL)."""
    for r in range(1, max_row + 1):
        for c in range(1, 6):
            v = normalizar(ws.cell(r, c).value)
            if v == label:
                return r
    return None


def _ler_valor_prestador(ws, linha_inicio: int, linha_fim: int,
                          provider: str, col_proj: int, col_manu: int) -> dict:
    """
    Busca o prestador em [linha_inicio..linha_fim] e retorna
    {projeto: float, manutencao: float} ou None se não encontrar.
    """
    prov_norm = normalizar(provider)
    for r in range(linha_inicio, linha_fim + 1):
        # Verificar col B (principal) e col F (secundária para TERCEIROS)
        for check_col in range(1, 8):
            val = normalizar(ws.cell(r, check_col).value)
            if val == prov_norm:
                proj = ws.cell(r, col_proj).value
                manu = ws.cell(r, col_manu).value
                return {
                    "projeto":    float(proj) if isinstance(proj, (int, float)) else 0.0,
                    "manutencao": float(manu) if isinstance(manu, (int, float)) else 0.0,
                }
    return None


def _encontrar_secao_v2(ws, max_row: int = 120) -> Optional[dict]:
    """
    Localiza a seção de V2 (abaixo das seções principais).
    Retorna {
      'matriz': {provider: {projeto: float, manutencao: float}},
      'filial':  {provider: {projeto: float, manutencao: float}}
    }
    """
    resultado = {"MATRIZ": {}, "FILIAL": {}}
    secao_atual = None

    # Colunas do bloco V2 (encontrar linha header)
    v2_proj_col = None
    v2_manu_col = None
    v2_prov_col = None
    v2_inicio   = None

    # Varrer o documento procurando pela seção V2
    # Identificada por ter 'PROJETO' e 'MANU' (sem a linha de cabeçalho principal)
    for r in range(20, max_row + 1):
        row = [normalizar(ws.cell(r, c).value) for c in range(1, 12)]

        # Detectar linha de header do V2
        if "PROJETO" in row and ("MANU" in row or "MANUTENÇÃO" in row):
            row_str = " ".join(row)
            # Verificar que é um header secundário (não é o header principal da linha 4)
            # Checando se está antes de uma linha com 'PRESTADOR' à esquerda
            has_prestador_col = "PRESTADOR" in row
            # O header V2 não tem "PRESTADOR" na mesma linha às vezes
            if r > 25:  # Deve estar depois das seções principais
                for c in range(1, 12):
                    v = normalizar(ws.cell(r, c).value)
                    if v == "PROJETO" and v2_proj_col is None:
                        v2_proj_col = c
                    if v in ("MANU", "MANUTENÇÃO") and v2_manu_col is None:
                        v2_manu_col = c
                    if v == "PRESTADOR":
                        v2_prov_col = c
                v2_inicio = r + 1
                break

        # Detectar seção matriz/filial no bloco V2
        for c in range(1, 8):
            v = normalizar(ws.cell(r, c).value)
            if v == "MATRIZ" and r > 25:
                secao_atual = "MATRIZ"
            elif v == "FILIAL" and r > 25:
                secao_atual = "FILIAL"

    if v2_inicio is None or v2_proj_col is None:
        return resultado

    # Ler os valores
    secao_atual = "MATRIZ"
    for r in range(v2_inicio, min(v2_inicio + 30, max_row + 1)):
        # Detectar mudança de seção
        for c in range(1, 8):
            v = normalizar(ws.cell(r, c).value)
            if v == "MATRIZ":
                secao_atual = "MATRIZ"
            elif v == "FILIAL":
                secao_atual = "FILIAL"

        # Tentar ler o nome do prestador
        prov = None
        # Buscar em col I (col 9) ou outras colunas
        for check_c in [9, 1, 2, v2_prov_col]:
            if check_c is None:
                continue
            v = normalizar(ws.cell(r, check_c).value)
            if (v and len(v) > 2
                    and v not in ("TOTAL", "MATRIZ", "FILIAL", "PRESTADOR",
                                  "PROJETO", "MANU", "MANUTENÇÃO", "")
                    and not v.replace(".", "").replace(",", "").isdigit()):
                prov = v
                break

        if prov:
            proj_v = ws.cell(r, v2_proj_col).value
            manu_v = ws.cell(r, v2_manu_col).value if v2_manu_col else None
            if isinstance(proj_v, (int, float)) or isinstance(manu_v, (int, float)):
                resultado[secao_atual][prov] = {
                    "projeto":    float(proj_v)  if isinstance(proj_v,  (int, float)) else 0.0,
                    "manutencao": float(manu_v)  if isinstance(manu_v,  (int, float)) else 0.0,
                }

    return resultado


def ler_referencia(ref_path: str, folder_name: str,
                   provider: str, frente: str) -> dict:
    """
    Lê os valores de referência V1/V2/V3 da planilha IMPOSTO para um
    prestador específico em uma frente específica.

    Retorna:
    {
      "aba":     str | None,
      "V1":      {"projeto": float, "manutencao": float} | None,
      "V2":      {"projeto": float, "manutencao": float} | None,
      "V3":      {"projeto": float, "manutencao": float} | None,
      "erro":    str | None,
    }
    """
    res = {"aba": None, "V1": None, "V2": None, "V3": None, "erro": None}

    try:
        wb = openpyxl.load_workbook(ref_path, read_only=False, data_only=True)
    except Exception as e:
        res["erro"] = f"Erro ao abrir referência: {e}"
        return res

    # Encontrar aba do mês
    aba = encontrar_aba_mes(wb, folder_name)
    if not aba:
        res["erro"] = f"Aba do mês '{folder_name}' não encontrada na planilha de referência."
        wb.close()
        return res

    res["aba"] = aba
    ws = wb[aba]

    # Detectar colunas
    cols = _detectar_colunas(ws)

    # Determinar profile e tipo
    is_matriz  = "MATRIZ" in frente.upper()
    is_projeto = "PROJETO" in frente.upper()
    profile    = "MATRIZ" if is_matriz else "FILIAL"

    # Encontrar linhas de início/fim de MATRIZ e FILIAL
    linha_matriz_inicio = _encontrar_secao(ws, "MATRIZ")
    linha_filial_inicio = _encontrar_secao(ws, "FILIAL")

    if not linha_matriz_inicio:
        res["erro"] = "Seção MATRIZ não encontrada na aba."
        wb.close()
        return res

    # Definir faixa de linhas para o profile
    if is_matriz:
        l_ini = linha_matriz_inicio + 1
        l_fim = (linha_filial_inicio - 1) if linha_filial_inicio else linha_matriz_inicio + 15
    else:
        if not linha_filial_inicio:
            res["erro"] = "Seção FILIAL não encontrada na aba."
            wb.close()
            return res
        l_ini = linha_filial_inicio + 1
        l_fim = linha_filial_inicio + 15

    # V3 = VALORES MEGALINK
    if cols["v3_proj"] and cols["v3_manu"]:
        dados_v3 = _ler_valor_prestador(ws, l_ini, l_fim, provider,
                                         cols["v3_proj"], cols["v3_manu"])
        res["V3"] = dados_v3

    # V1 = VALORES TERCEIROS
    if cols["v1_proj"] and cols["v1_manu"]:
        dados_v1 = _ler_valor_prestador(ws, l_ini, l_fim, provider,
                                         cols["v1_proj"], cols["v1_manu"])
        res["V1"] = dados_v1

    # V2 = seção abaixo (tabela secundária)
    v2_data = _encontrar_secao_v2(ws)
    prov_norm = normalizar(provider)
    if prov_norm in v2_data.get(profile, {}):
        res["V2"] = v2_data[profile][prov_norm]
    elif v2_data.get(profile):
        # Tentar correspondência parcial
        for k, v in v2_data[profile].items():
            if prov_norm in k or k in prov_norm:
                res["V2"] = v
                break

    wb.close()
    return res


def comparar_com_referencia(ref_path: str, folder_name: str,
                             provider: str, result_validacao: dict) -> dict:
    """
    Para cada frente × versão, compara TOTAL_GERAL do arquivo com
    valor de referência (PROJETO + MANUTENÇÃO) da planilha IMPOSTO.

    Retorna:
    {
      frente: {
        "referencia_ok": bool,
        "comparacoes": [
          {"versao": "V1", "total_arquivo": float, "referencia": float,
           "match": bool, "diferenca": float}
        ]
      }
    }
    """
    comparacoes = {}

    for frente, fd in result_validacao.get("frentes", {}).items():
        ref = ler_referencia(ref_path, folder_name, provider, frente)
        comps = []
        tudo_ok = True

        for ver in ["V1", "V2", "V3"]:
            vd = fd.get(ver) or {}
            total_arq = vd.get("total_geral", 0.0)
            ref_vals  = ref.get(ver)
            ref_aba   = ref.get("aba")
            ref_erro  = ref.get("erro")

            if ref_vals:
                is_proj = "PROJETO" in frente.upper()
                ref_val = ref_vals["projeto"] if is_proj else ref_vals["manutencao"]
                dif     = round(abs(total_arq - ref_val), 2)
                match   = dif <= 0.02
                if not match:
                    tudo_ok = False
                comps.append({
                    "versao":       ver,
                    "total_arquivo": total_arq,
                    "referencia":   ref_val,
                    "match":        match,
                    "diferenca":    dif,
                    "aba":          ref_aba,
                    "erro":         None,
                })
            else:
                comps.append({
                    "versao":       ver,
                    "total_arquivo": total_arq,
                    "referencia":   None,
                    "match":        None,
                    "diferenca":    None,
                    "aba":          ref_aba,
                    "erro":         ref_erro or "Valor não encontrado na referência",
                })

        comparacoes[frente] = {
            "referencia_ok": tudo_ok,
            "aba_referencia": ref.get("aba"),
            "comparacoes": comps,
        }

    return comparacoes
